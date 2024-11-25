import yfinance as yf
import pandas as pd
import talib
import numpy as np
from ModelDataPreprocess import preprocess_single_stock
import tensorflow as tf   
from openpyxl import Workbook
import requests
from bs4 import BeautifulSoup
from tqdm import tqdm

# 股票買入與賣出條件
def analyze_stock(stock):
    ticker = yf.Ticker(stock)
    df = ticker.history(period="3mo")
    if df.empty:
        print(f"Stock {stock} data download failed or no data available.")
        return None
    if df['Close'].values.reshape(-1, 1).astype('float32').shape[0] < 30:
        print(f"Stock {stock} data length less than 30.")
        return None

    # 計算指標
    df['RSI'] = talib.RSI(df['Close'], timeperiod=14)
    df['RSI7'] = talib.RSI(df['Close'], timeperiod=7)
    df['MACD'], df['MACDSignal'], df['MACDHist'] = talib.MACD(df['Close'], fastperiod=12, slowperiod=26, signalperiod=9)

    # 檢查初步賣出條件
    last_rsi = df['RSI'].iloc[-1]
    last_rsi7 = df['RSI7'].iloc[-1]
    last_macd = df['MACD'].iloc[-1]
    last_macd_signal = df['MACDSignal'].iloc[-1]

    # 直接判斷賣出條件
    if (last_rsi > 70 and (last_macd < last_macd_signal or last_rsi7 < last_rsi)):
    # if (last_rsi > 70):
        return df, "sell"

    # 檢查初步買入條件
    if (last_rsi < 30 and (last_macd > last_macd_signal or last_rsi7 > last_rsi)):
    # if (last_rsi < 30):
        return df, "buy"

    # 不符合买入或卖出条件的股票
    return df, "watch"
    

# 使用模型進行預測
def model_prediction(df):
    X, y, scaler = preprocess_single_stock(df['Close'])  # 預處理數據
    loaded_model = tf.keras.models.load_model("stock_prediction_model.h5")
    # 只取最後一筆數據進行預測
    last_X = X[-1].reshape(1, X.shape[1], 1)
    predicted_close = loaded_model.predict(last_X)
    predicted_close = scaler.inverse_transform(predicted_close)[0][0]
    last_close = df['Close'].iloc[-1]
    return predicted_close, last_close

# 保存交易信號到 Excel
def save_trade_signals_to_excel(stock_list):
    wb_buy = Workbook()
    ws_buy = wb_buy.active
    ws_buy.append(['Stock', 'Date', 'Close', 'Predicted Close', 'RSI', 'RSI7 > RSI14', 'MACD > MACDSignal', 'Buy Signal'])
    # ws_buy.append(['Stock', 'Date', 'Close', 'Predicted Close', 'RSI', 'RSI7', 'MACD', 'MACDSignal', 'Buy Signal'])

    wb_sell = Workbook()
    ws_sell = wb_sell.active
    ws_sell.append(['Stock', 'Date', 'Close', 'RSI', 'RSI7 < RSI14', 'MACD < MACDSignal', 'Sell Signal'])
    # ws_sell.append(['Stock', 'Date', 'Close', 'RSI', 'RSI7', 'MACD', 'MACDSignal', 'Sell Signal'])

    wb_watch = Workbook()
    ws_watch = wb_watch.active
    ws_watch.append(['Stock', 'Date', 'Close', 'RSI', 'RSI7 > RSI14', 'MACD > MACDSignal'])
    # ws_watch.append(['Stock', 'Date', 'Close', 'RSI', 'RSI7', 'MACD', 'MACDSignal'])

    for stock in tqdm(stock_list, desc="Processing stocks"):
        analysis_result = analyze_stock(stock)
        if analysis_result is not None:
            df, signal_type = analysis_result
            print(f'{stock} signal_type:', signal_type)

            if signal_type == "buy":
                predicted_close, last_close = model_prediction(df)
                # 添加預測的收盤價
                df['PredictedClose'] = predicted_close
                # 買入信號
                buy_signal = ((df['RSI'] < 30) & ((df['MACD'] > df['MACDSignal']) | (predicted_close > last_close)) | (df['RSI7'] > df['RSI']))
                # buy_signal = (df['RSI'] < 30)
                df['BuySignal'] = buy_signal
                last_buy_signal = df[df['BuySignal']].iloc[-1]
                ws_buy.append([
                    stock,
                    last_buy_signal.name.replace(tzinfo=None),
                    last_buy_signal['Close'],
                    last_buy_signal['PredictedClose'],
                    last_buy_signal['RSI'],
                    last_buy_signal['RSI7'] > last_buy_signal['RSI'],
                    last_buy_signal['MACD'] > last_buy_signal['MACDSignal'],
                    last_buy_signal['BuySignal']
                ])
                # buy_signals = df[df['BuySignal']]
                # for index, row in buy_signals.iterrows():
                #     ws_buy.append([
                #         stock,
                #         row.name.replace(tzinfo=None),
                #         row['Close'],
                #         row['PredictedClose'],
                #         row['RSI'],
                #         row['RSI7'] > row['RSI'],
                #         row['MACD'] > row['MACDSignal'],
                #         row['BuySignal']
                #     ])

            elif signal_type == "sell":
                # 賣出信號
                sell_signal = ((df['RSI'] > 70) & ((df['MACD'] < df['MACDSignal'])) | (df['RSI7'] < df['RSI']))
                # sell_signal = (df['RSI'] > 70)
                df['SellSignal'] = sell_signal
                last_sell_signal = df[df['SellSignal']].iloc[-1]
                ws_sell.append([
                    stock,
                    last_sell_signal.name.replace(tzinfo=None),
                    last_sell_signal['Close'],
                    last_sell_signal['RSI'],
                    last_sell_signal['RSI7'] < last_sell_signal['RSI'],
                    last_sell_signal['MACD'] < last_sell_signal['MACDSignal'],
                    last_sell_signal['SellSignal']
                ])
                # sell_signals = df[df['SellSignal']]
                # for index, row in sell_signals.iterrows():
                #     ws_sell.append([
                #         stock,
                #         row.name.replace(tzinfo=None),
                #         row['Close'],
                #         row['RSI'],
                #         row['RSI7'] < row['RSI'],
                #         row['MACD'] < row['MACDSignal'],
                #         row['SellSignal']
                #     ])

            else:
                # 觀察信號
                last_watch_signal = df.iloc[-1]
                ws_watch.append([
                    stock,
                    last_watch_signal.name.replace(tzinfo=None),
                    last_watch_signal['Close'],
                    last_watch_signal['RSI'],
                    last_watch_signal['RSI7'] > last_watch_signal['RSI'],
                    last_watch_signal['MACD'] > last_watch_signal['MACDSignal']
                ])
                # for index, row in df.iterrows():
                #     ws_watch.append([
                #         stock,
                #         row.name.replace(tzinfo=None),
                #         row['Close'],
                #         row['RSI'],
                #         row['RSI7'] > row['RSI'],
                #         row['MACD'] > row['MACDSignal']
                #     ])

    wb_buy.save("buy_signals.xlsx")
    wb_sell.save("sell_signals.xlsx")
    wb_watch.save("watch_signals.xlsx")

def download_stock_list():
    url = "https://isin.twse.com.tw/isin/C_public.jsp?strMode=2"
    res = requests.get(url)
    soup = BeautifulSoup(res.text, "lxml")
    tr = soup.findAll('tr')

    tds = []
    for raw in tr:
        data = [td.get_text() for td in raw.findAll("td")]
        if len(data) == 7:
            tds.append(data)

    df = pd.DataFrame(tds[1:], columns=tds[0])

    second_column = df.iloc[:, 0]
    second_column_split = second_column.str.split('　', expand=True)
    second_column_split.columns = ['Stock Code', 'Company Name']
    df = second_column_split.iloc[:997]

    stock_list = df.iloc[:, 0].astype(str).tolist()
    stock_name = df.iloc[:, 1].astype(str).tolist()
    stock_no = [stock_id + '.TW' for stock_id in stock_list]
    return stock_no

# 檢查多支股票的交易信號
stock_list = download_stock_list()
# stock_list = ['2330.TW']
save_trade_signals_to_excel(stock_list)
print("Finished processing and saved to Excel files.")